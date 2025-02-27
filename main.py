import os
import pandas as pd
import shutil
import openpyxl
import subprocess

# basic_text = '''0
# up
# p_inE+06
# p_outE+06
# 16
# 30
# 1
# 2.127
# 0.023
# 0.032
# 0.366
# 4.000000E-03
# 4.737000E-03
# 1.000000E-03
# 4.200000E-03
# 3.200000E-03
# 4.200000E-03
# 1.000000E-03
# 12
# 1
# 0.013893
# 0.75000E-03
# 0.75000E-03
# 1.4
# 287.1
# T_in
# 5.77
# 8.0
# 1.00000E-03
# 1.00000E-03
# 0.023
# a
# p
# 0.00001
# 0.00001
# 0.00002
# 0.00001
# 0.00'''

var_name = input("Enter var file name without extension\n")

with open(f"{var_name}.dat", 'r') as var_file:
    var_text = var_file.read()

var_text_split = var_text.split('\n')

data = pd.read_excel("ГУ для расчета.xlsx", header=1)

try:
    os.mkdir(var_name)
except FileExistsError:
    shutil.rmtree(var_name)
    os.mkdir(var_name)

for i, row in data.iterrows():
    if row['n'] == 0:
        up_limit = 1
    elif row['n'] < 1000:
        up_limit = 1000
    else:
        up_limit = 22000

    var_text_split[0] = '0'
    var_text_split[1] = str(up_limit)
    var_text_split[2] = str(row['pвх*']) + 'E+06'
    var_text_split[3] = str(row['pвых']) + 'E+06'
    var_text_split[25] = str(row['Tвх*'])

    new_text = '\n'.join(var_text_split)

    # new_text = (basic_text
    #             .replace('up', str(up_limit))
    #             .replace('p_in', str(row['pвх*']))
    #             .replace('p_out', str(row['pвых']))
    #             .replace('T_in', str(row['Tвх*'])))

    dir_name = os.path.join(var_name, str(row['t']).ljust(7, '0'))

    os.mkdir(dir_name)

    with open(os.path.join(dir_name, 'var.dat'), 'w') as var_file:
        var_file.write(new_text)

    abs_exe_path = os.path.abspath('app_with_sas.exe')
    subprocess.run(abs_exe_path, cwd=os.path.abspath(dir_name))

    with open(os.path.join(dir_name, 'res.txt'), 'r') as res_file:
        res_split = res_file.read().replace('.', ',').split()

    pass

    wb = openpyxl.load_workbook('Форма обработки одномерного расчета.xlsx', keep_vba=False)

    ws = wb['Сокращенная форма']
    ws['B1'] = row['PiT']
    ws['B4'] = row['Tвх*']

    ws = wb['Графики']
    ws['A3'] = 't, с'
    ws['A4'] = row['t']

    ws['A6'] = 'p_in, МПа'
    ws['A7'] = row['pвх*']
    ws['B6'] = 'p_out, МПа'
    ws['B7'] = row['pвых']

    ws = wb['Лист0']
    letters = ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I')
    for ind, res_data in enumerate(res_split):
        letter = letters[ind % len(letters)]
        row = ind // len(letters) + 1
        row += (ind // 144) * 2
        ws[letter + str(row)] = res_data

    wb.save(os.path.join(dir_name, "Форма обработки одномерного расчета.xlsx"))
    wb.close()

